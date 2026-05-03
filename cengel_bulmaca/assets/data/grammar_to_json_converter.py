#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel to JSON Converter for Grammar Questions (Dil Bilgisi Soruları)
Grammar_clues.json dosyası için özel dönüştürücü
Excel formatından dilBilgisiSorulari.xlsx dosyasını JSON'a çevirir
"""

import json
import re
import os
import sys
import openpyxl
from typing import Dict, List

# UTF-8 encoding'i ayarla
if sys.stdout.encoding != 'utf-8':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


def clean_text(text: str) -> str:
    """Metni temizle ve normalize et"""
    if not text:
        return ""
    text = str(text).strip()
    # Fazla boşlukları kaldır
    text = re.sub(r'\s+', ' ', text)
    return text


def turkish_upper(text: str) -> str:
    """Türkçe karakterleri duyarlı bir şekilde büyüt"""
    # Türkçe karakter eşlemeleri
    turkish_map = {
        'i': 'İ', 'ı': 'I',
        'a': 'A', 'b': 'B', 'c': 'C', 'ç': 'Ç', 'd': 'D', 'e': 'E',
        'f': 'F', 'g': 'G', 'ğ': 'Ğ', 'h': 'H', 'j': 'J', 'k': 'K',
        'l': 'L', 'm': 'M', 'n': 'N', 'o': 'O', 'ö': 'Ö', 'p': 'P',
        'r': 'R', 's': 'S', 'ş': 'Ş', 't': 'T', 'u': 'U', 'ü': 'Ü',
        'v': 'V', 'y': 'Y', 'z': 'Z', 'x': 'X', 'w': 'W', 'q': 'Q'
    }
    result = []
    for char in text:
        result.append(turkish_map.get(char, char.upper()))
    return ''.join(result)


def clean_answer(answer: str) -> str:
    """Cevabı çengel bulmaca için temizle (Türkçe karakterleri koru)"""
    if not answer:
        return ""
    answer = str(answer).strip()
    
    # Türkçe karakterleri duyarlı bir şekilde büyüt
    answer = turkish_upper(answer)
    # Boşlukları kaldır ama Türkçe karakterleri koru
    answer = re.sub(r'\s+', '', answer)
    # Türkçe ve İngilizce harfler dışındaki karakterleri kaldır
    answer = re.sub(r'[^a-zA-ZÇçĞğİıÖöŞşÜü]', '', answer)
    return answer


def get_difficulty_value(difficulty_text: str) -> int:
    """Zorluk metnini sayısal değere çevir"""
    if not difficulty_text:
        return 2
    diff = str(difficulty_text).strip().upper()
    if 'KOLAY' in diff or diff == '1':
        return 1
    elif 'ZOR' in diff or diff == '3':
        return 3
    return 2  # ORTA veya varsayılan


def create_category_id(name: str) -> str:
    """Kategori adından ID oluştur"""
    # Türkçe karakterleri değiştir
    id_str = name.lower()
    replacements = {
        'ı': 'i', 'ğ': 'g', 'ü': 'u', 'ş': 's', 'ö': 'o', 'ç': 'c',
        'İ': 'i', 'Ğ': 'g', 'Ü': 'u', 'Ş': 's', 'Ö': 'o', 'Ç': 'c',
        ' ': '_', '-': '_', '.': '', "'": '', '"': ''
    }
    for old, new in replacements.items():
        id_str = id_str.replace(old, new)
    id_str = re.sub(r'[^a-z0-9_]', '', id_str)
    id_str = re.sub(r'_+', '_', id_str)
    return id_str.strip('_')


def parse_excel_sheet(ws, sheet_name: str) -> Dict:
    """Bir Excel sayfasını parse et ve kategori oluştur"""
    category_id = create_category_id(sheet_name)
    
    clues = []
    clue_id = 0
    
    # Başlık satırını oku
    header_row = 1
    headers = []
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers.append(str(cell.value).strip().upper())
        else:
            headers.append(f"COL_{col_idx}")
    
    # Sütun indekslerini bul
    soru_col = None
    cevap_col = None
    zorluk_col = None
    
    for idx, header in enumerate(headers):
        if 'SORU' in header or 'QUESTION' in header:
            soru_col = idx
        elif 'CEVAP' in header or 'ANSWER' in header:
            cevap_col = idx
        elif 'ZORLUK' in header or 'PUAN' in header or 'DIFFICULTY' in header:
            zorluk_col = idx
    
    # Eğer başlık bulunamadıysa varsayılan değerleri kullan
    if soru_col is None:
        soru_col = 0  # İlk sütun
    if cevap_col is None:
        cevap_col = 1  # İkinci sütun
    if zorluk_col is None:
        zorluk_col = 2  # Üçüncü sütun
    
    print(f"  Sütun eşleşmeleri - Soru: {soru_col}, Cevap: {cevap_col}, Zorluk: {zorluk_col}")
    
    # Veri satırlarını oku
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or len(row) < 2:
            continue
            
        # Eğer satır yeterli sütun içermiyorsa atla
        if len(row) <= max(soru_col, cevap_col):
            continue
            
        question = row[soru_col] if soru_col < len(row) else None
        answer = row[cevap_col] if cevap_col < len(row) else None
        difficulty_text = row[zorluk_col] if zorluk_col < len(row) else None
        
        # Boş veya başlık satırlarını atla
        if not question or not answer:
            continue
            
        question_str = str(question).strip()
        if question_str.upper() in ['SORU', 'SINIF', 'QUESTION', '']:
            continue
        
        # Cevabı temizle
        cleaned_answer = clean_answer(answer)
        
        # Çok kısa veya çok uzun cevapları atla
        if len(cleaned_answer) < 2 or len(cleaned_answer) > 30:
            print(f"    [!] Atlanan (uzunluk): {answer} -> {cleaned_answer}")
            continue
        
        # Geçerli soru/cevap ekle
        clue_id += 1
        clues.append({
            "id": f"{category_id}_{clue_id}",
            "question": clean_text(question),
            "answer": cleaned_answer,
            "difficulty": get_difficulty_value(difficulty_text)
        })
    
    return {
        "id": category_id,
        "name": sheet_name,
        "clues": clues
    }


def process_grammar_excel(excel_path: str) -> Dict:
    """Dil bilgisi Excel dosyasını işle ve JSON oluştur"""
    categories = []
    
    if not os.path.exists(excel_path):
        print(f"[ERROR] Dosya bulunamadı: {excel_path}")
        return None
    
    print(f"\n[*] İşleniyor: {os.path.basename(excel_path)}")
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  [+] Sayfa: {sheet_name}")
            
            category = parse_excel_sheet(ws, sheet_name)
            
            if category["clues"]:
                categories.append(category)
                print(f"     [OK] {len(category['clues'])} soru/cevap bulundu")
            else:
                print(f"     [!] Boş sayfa atlandı")
                
    except Exception as e:
        print(f"  [ERROR] Dosya okunurken hata: {e}")
        return None
    
    # JSON çıktısı oluştur
    output_data = {
        "version": "1.0",
        "description": "Dil bilgisi çengel bulmaca verileri - kategorilere ayrılmış",
        "categories": categories
    }
    
    return output_data


def main():
    # Yolları belirle
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.abspath(os.path.join(script_dir, '..', '..', '..', 'bulmacalar', 'dilBilgisiSorulari.xlsx'))
    output_path = os.path.join(script_dir, 'grammar_clues.json')
    
    print("=" * 70)
    print("📚 Excel -> JSON Dönüştürücü (Dil Bilgisi Soruları)")
    print("=" * 70)
    print(f"Excel Dosyası: {excel_path}")
    print(f"Çıktı Dosyası: {output_path}")
    print("=" * 70)
    
    # Dönüştür
    data = process_grammar_excel(excel_path)
    
    if data is None:
        print("\n[ERROR] Dönüştürme başarısız!")
        return
    
    # JSON dosyasına yaz
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        total_clues = sum(len(c["clues"]) for c in data['categories'])
        
        print("\n" + "=" * 70)
        print("[✓] Dönüştürme başarıyla tamamlandı!")
        print(f"[📊] Toplam {len(data['categories'])} kategori")
        print(f"[📊] Toplam {total_clues} soru/cevap")
        print(f"[📁] Çıktı: {output_path}")
        print("=" * 70)
        
    except Exception as e:
        print(f"\n[ERROR] JSON dosyası yazılırken hata: {e}")


if __name__ == "__main__":
    main()
