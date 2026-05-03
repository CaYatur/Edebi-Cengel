#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel to JSON Converter for Crossword Puzzles
Excel dosyalarından çengel bulmaca verilerini JSON formatına dönüştürür
"""

import json
import re
import os
import sys
import openpyxl
from typing import Dict, List, Tuple

# UTF-8 encoding'i ayarla
if sys.stdout.encoding != 'utf-8':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


def clean_text(text: str) -> str:
    """Metni temizle ve normalize et"""
    if not text:
        return ""
    text = str(text).strip()
    # Fazla boşlukları kaldır
    text = re.sub(r'\s+', ' ', text)
    return text


def turkish_upper(text: str) -> str:
    """Türkçe karakterleri duyarlı bir şekilde büyüt"""
    # Türkçe karakter eşlemeleri
    turkish_map = {
        'i': 'İ', 'ı': 'I',
        'a': 'A', 'b': 'B', 'c': 'C', 'ç': 'Ç', 'd': 'D', 'e': 'E',
        'f': 'F', 'g': 'G', 'ğ': 'Ğ', 'h': 'H', 'j': 'J', 'k': 'K',
        'l': 'L', 'm': 'M', 'n': 'N', 'o': 'O', 'ö': 'Ö', 'p': 'P',
        'r': 'R', 's': 'S', 'ş': 'Ş', 't': 'T', 'u': 'U', 'ü': 'Ü',
        'v': 'V', 'y': 'Y', 'z': 'Z', 'x': 'X', 'w': 'W', 'q': 'Q'
    }
    result = []
    for char in text:
        result.append(turkish_map.get(char, char.upper()))
    return ''.join(result)


def clean_answer(answer: str) -> str:
    """Cevabı çengel bulmaca için temizle (Türkçe karakterleri koru)"""
    if not answer:
        return ""
    answer = str(answer).strip()
    
    # Türkçe karakterleri duyarlı bir şekilde büyüt
    answer = turkish_upper(answer)
    # Boşlukları kaldır ama Türkçe karakterleri koru
    answer = re.sub(r'\s+', '', answer)
    # Türkçe ve İngilizce harfler dışındaki karakterleri kaldır
    # A-Z, Ç, Ğ, İ, Ö, Ş, Ü (ve lowercase versiyonları)
    answer = re.sub(r'[^a-zA-ZÇçĞğİıÖöŞşÜü]', '', answer)
    return answer


def get_difficulty_value(difficulty_text: str) -> int:
    """Zorluk metnini sayısal değere çevir"""
    if not difficulty_text:
        return 2
    diff = str(difficulty_text).strip().upper()
    if 'KOLAY' in diff:
        return 1
    elif 'ZOR' in diff:
        return 3
    return 2  # ORTA veya varsayılan


def create_category_id(name: str) -> str:
    """Kategori adından ID oluştur"""
    # Türkçe karakterleri değiştir
    id_str = name.lower()
    replacements = {
        'ı': 'i', 'ğ': 'g', 'ü': 'u', 'ş': 's', 'ö': 'o', 'ç': 'c',
        'İ': 'i', 'Ğ': 'g', 'Ü': 'u', 'Ş': 's', 'Ö': 'o', 'Ç': 'c',
        ' ': '_', '-': '_', '.': '', "'": '', '"': ''
    }
    for old, new in replacements.items():
        id_str = id_str.replace(old, new)
    id_str = re.sub(r'[^a-z0-9_]', '', id_str)
    id_str = re.sub(r'_+', '_', id_str)
    return id_str.strip('_')


def parse_excel_sheet(ws, sheet_name: str) -> Dict:
    """Bir Excel sayfasını parse et"""
    category_id = create_category_id(sheet_name)
    
    clues = []
    clue_id = 0
    
    # Başlık satırını bul
    header_row = 1
    headers = []
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers.append(str(cell.value).strip().upper())
        else:
            headers.append(f"COL_{col_idx}")
    
    # Sütun indekslerini bul
    soru_col = None
    cevap_col = None
    zorluk_col = None
    
    for idx, header in enumerate(headers):
        if 'SORU' in header:
            soru_col = idx
        elif 'CEVAP' in header:
            cevap_col = idx
        elif 'ZORLUK' in header or 'PUAN' in header:
            zorluk_col = idx
    
    # Eğer başlık bulunamadıysa varsayılan değerleri kullan
    if soru_col is None:
        soru_col = 1 if len(headers) > 1 else 0
    if cevap_col is None:
        cevap_col = 2 if len(headers) > 2 else 1
    if zorluk_col is None:
        zorluk_col = 3 if len(headers) > 3 else 2
    
    # Veri satırlarını oku
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or len(row) <= max(soru_col, cevap_col):
            continue
            
        question = row[soru_col] if soru_col < len(row) else None
        answer = row[cevap_col] if cevap_col < len(row) else None
        difficulty_text = row[zorluk_col] if zorluk_col < len(row) else None
        
        # Boş veya başlık satırlarını atla
        if not question or not answer:
            continue
        question_str = str(question).strip()
        if question_str.upper() in ['SORU', 'SINIF', '']:
            continue
            
        cleaned_answer = clean_answer(answer)
        if len(cleaned_answer) < 2 or len(cleaned_answer) > 20:
            print(f"  Atlanan (uzunluk): {answer} -> {cleaned_answer}")
            continue
            
        clue_id += 1
        clues.append({
            "id": f"{category_id}_{clue_id}",
            "question": clean_text(question),
            "answer": cleaned_answer,
            "difficulty": get_difficulty_value(difficulty_text)
        })
    
    return {
        "id": category_id,
        "name": sheet_name,
        "clues": clues
    }


def process_excel_files(excel_dir: str, output_dir: str) -> Dict:
    """Excel dosyalarını işle ve JSON oluştur"""
    categories = []
    
    # Excel dosyalarını bul
    excel_files = [f for f in os.listdir(excel_dir) if f.endswith('.xlsx') and not f.startswith('~$')]
    
    for excel_file in sorted(excel_files):
        filepath = os.path.join(excel_dir, excel_file)
        print(f"\n[*] Isleniyor: {excel_file}")
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"  [+] Sayfa: {sheet_name}")
                
                category = parse_excel_sheet(ws, sheet_name)
                
                if category["clues"]:
                    categories.append(category)
                    print(f"     [OK] {len(category['clues'])} soru/cevap bulundu")
                else:
                    print(f"     [!] Bos sayfa atlanmadi")
                    
        except Exception as e:
            print(f"  [ERROR] Hata: {e}")
    
    # JSON çıktısı oluştur
    output_data = {
        "version": "2.0",
        "description": "Kategorilere ayrılmış çengel bulmaca verileri - dinamik oluşturma için",
        "categories": categories,
        "totalClues": sum(len(c["clues"]) for c in categories)
    }
    
    return output_data


def main():
    # Yolları belirle
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_dir = os.path.abspath(os.path.join(script_dir, '..', '..', '..', 'bulmacalar'))
    output_dir = script_dir
    
    print("=" * 60)
    print("🔄 Excel -> JSON Dönüştürücü")
    print("=" * 60)
    print(f"Excel dizini: {excel_dir}")
    print(f"Çıktı dizini: {output_dir}")
    
    # Dönüştür
    data = process_excel_files(excel_dir, output_dir)
    
    # JSON dosyasına yaz
    output_path = os.path.join(output_dir, 'crossword_clues.json')
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print("\n" + "=" * 60)
    print(f"[SUCCESS] Donusturme tamamlandi!")
    print(f"[STATS] Toplam {len(data['categories'])} kategori")
    print(f"[STATS] Toplam {data['totalClues']} soru/cevap")
    print(f"[OUTPUT] Cikti: {output_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
